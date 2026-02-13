"""
Questionnaire Mapper - Maps evidence to questionnaire questions
"""
from typing import Dict, List, Any
import json
import openpyxl
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np


class QuestionnaireMapper:
    """Map extracted evidence to questionnaire questions"""

    def __init__(self):
        # Load sentence transformer for semantic similarity
        try:
            self.model = SentenceTransformer('all-MiniLM-L6-v2')
        except:
            self.model = None
        self.question_mappings = []

    def load_questionnaire(self, excel_path: str) -> List[Dict[str, Any]]:
        """Load questions from Datadog questionnaire"""
        questions = []

        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # Try to find the questions sheet (common names)
            sheet_names = ['Questions', 'Questionnaire', 'Assessment', workbook.sheetnames[0]]
            sheet = None

            for name in sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    break

            if not sheet:
                return questions

            # Parse questions (assuming structure: ID, Category, Question, ...)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not row or not any(row):
                    continue

                # Try to extract question (usually in column 2 or 3)
                question_text = None
                category = None

                for cell in row:
                    if cell and isinstance(cell, str) and len(cell) > 20 and '?' in cell:
                        question_text = cell
                        break

                if question_text:
                    questions.append({
                        "id": f"Q{row_idx}",
                        "question": question_text,
                        "category": category or "General",
                        "row_num": row_idx
                    })

        except Exception as e:
            print(f"Error loading questionnaire: {e}")

        return questions

    def calculate_similarity(self, question: str, evidence_text: str) -> float:
        """Calculate semantic similarity between question and evidence"""
        if not self.model:
            # Fallback to keyword matching
            question_words = set(question.lower().split())
            evidence_words = set(evidence_text.lower().split())
            overlap = len(question_words & evidence_words)
            return overlap / len(question_words) if question_words else 0.0

        try:
            # Use sentence transformer
            embeddings = self.model.encode([question, evidence_text])
            similarity = cosine_similarity([embeddings[0]], [embeddings[1]])[0][0]
            return float(similarity)
        except:
            return 0.0

    def map_evidence_to_questions(
        self,
        questions: List[Dict[str, Any]],
        evidence_library: List[Dict[str, Any]],
        threshold: float = 0.3
    ) -> List[Dict[str, Any]]:
        """Map evidence to each question"""
        self.question_mappings = []

        for question in questions:
            question_text = question['question']
            matched_evidence = []

            # Find relevant evidence
            for evidence in evidence_library:
                similarity = self.calculate_similarity(question_text, evidence['text'])

                if similarity >= threshold:
                    matched_evidence.append({
                        "evidence_text": evidence['text'],
                        "source": evidence['source'],
                        "keywords": evidence.get('keywords', []),
                        "similarity_score": similarity,
                        "evidence_type": evidence.get('type', 'unknown')
                    })

            # Sort by similarity
            matched_evidence.sort(key=lambda x: x['similarity_score'], reverse=True)

            # Determine answer and confidence
            answer, confidence = self._generate_answer(matched_evidence)
            gaps = self._identify_gaps(question_text, matched_evidence)

            self.question_mappings.append({
                "question_id": question['id'],
                "question": question_text,
                "category": question.get('category', 'General'),
                "answer": answer,
                "evidence": matched_evidence[:5],  # Top 5 most relevant
                "confidence": confidence,
                "gaps": gaps
            })

        return self.question_mappings

    def _generate_answer(self, evidence_list: List[Dict[str, Any]]) -> tuple:
        """Generate answer and confidence level from evidence"""
        if not evidence_list:
            return "Insufficient Evidence", "NOT_FOUND"

        # High confidence if we have strong evidence
        if evidence_list[0]['similarity_score'] > 0.6:
            # Extract answer from top evidence
            top_evidence = evidence_list[0]['evidence_text']
            return f"Yes - {top_evidence[:200]}...", "HIGH"
        elif evidence_list[0]['similarity_score'] > 0.4:
            return "Partially Addressed - See evidence", "MEDIUM"
        else:
            return "Insufficient Evidence", "LOW"

    def _identify_gaps(self, question: str, evidence_list: List[Dict[str, Any]]) -> List[str]:
        """Identify gaps in evidence"""
        gaps = []

        if not evidence_list:
            gaps.append("No evidence found in vendor documentation")
        elif evidence_list[0]['similarity_score'] < 0.4:
            gaps.append("Evidence is weak or indirect")

        if len(evidence_list) < 2:
            gaps.append("Single source only - requires corroboration")

        return gaps

    def to_json(self, output_path: str = "questionnaire_mapping.json"):
        """Save mappings to JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.question_mappings, f, indent=2, ensure_ascii=False)
        return output_path

    def to_excel(self, output_path: str = "completed_questionnaire.xlsx"):
        """Export to Excel format"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Completed Assessment"

        # Headers
        headers = ["Question ID", "Category", "Question", "Answer", "Evidence References", "Confidence", "Gaps/Follow-ups"]
        sheet.append(headers)

        # Data rows
        for mapping in self.question_mappings:
            evidence_refs = "; ".join([e['source'] for e in mapping['evidence'][:3]])
            gaps = "; ".join(mapping['gaps']) if mapping['gaps'] else "None"

            sheet.append([
                mapping['question_id'],
                mapping['category'],
                mapping['question'],
                mapping['answer'],
                evidence_refs,
                mapping['confidence'],
                gaps
            ])

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(output_path)
        return output_path
