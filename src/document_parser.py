"""
Document Parser - Extracts text and tables from vendor documents
"""
import pypdf
import pdfplumber
import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
import json


class DocumentParser:
    """Parse vendor documents (PDF, Excel, Word)"""

    def __init__(self):
        self.parsed_docs = []

    def parse_pdf(self, file_path: str) -> Dict[str, Any]:
        """Extract text and tables from PDF"""
        result = {
            "type": "pdf",
            "filename": Path(file_path).name,
            "pages": [],
            "tables": []
        }

        try:
            # Extract text
            with open(file_path, 'rb') as file:
                pdf_reader = pypdf.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    text = page.extract_text()
                    result["pages"].append({
                        "page_num": page_num,
                        "text": text
                    })

            # Extract tables
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    tables = page.extract_tables()
                    if tables:
                        for table_idx, table in enumerate(tables):
                            result["tables"].append({
                                "page": page_num,
                                "table_index": table_idx,
                                "data": table
                            })
        except Exception as e:
            result["error"] = str(e)

        return result

    def parse_excel(self, file_path: str) -> Dict[str, Any]:
        """Extract data from Excel files"""
        result = {
            "type": "excel",
            "filename": Path(file_path).name,
            "sheets": []
        }

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Convert to list of lists
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data.append(list(row))

                result["sheets"].append({
                    "sheet_name": sheet_name,
                    "data": data,
                    "max_row": sheet.max_row,
                    "max_col": sheet.max_column
                })
        except Exception as e:
            result["error"] = str(e)

        return result

    def parse_document(self, file_path: str) -> Dict[str, Any]:
        """Auto-detect and parse document based on extension"""
        path = Path(file_path)
        extension = path.suffix.lower()

        if extension == '.pdf':
            return self.parse_pdf(file_path)
        elif extension in ['.xlsx', '.xls']:
            return self.parse_excel(file_path)
        else:
            return {
                "type": "unknown",
                "filename": path.name,
                "error": f"Unsupported file type: {extension}"
            }

    def parse_all(self, file_paths: List[str]) -> List[Dict[str, Any]]:
        """Parse multiple documents"""
        self.parsed_docs = []
        for file_path in file_paths:
            doc = self.parse_document(file_path)
            self.parsed_docs.append(doc)
        return self.parsed_docs

    def to_json(self, output_path: str = "parsed_documents.json"):
        """Save parsed documents to JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.parsed_docs, f, indent=2, ensure_ascii=False)
        return output_path
